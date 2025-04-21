import os
import shutil
import time
import traceback
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
import logging
import openpyxl
from collections import Counter
import re
import pytz
from pymongo import MongoClient
from pymongo.errors import ServerSelectionTimeoutError, AutoReconnect
import psutil
import win32com.client
import ssl

# Configure logging with more detail
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', handlers=[
    logging.FileHandler("debug.log"),
    logging.StreamHandler()
])

# Directory definitions
REPORT_DIR = r"C:\Users\ASUS\Desktop\New folder\Input"
REF_DIR = r"C:\Users\ASUS\Desktop\New folder\Reference"

# MongoDB setup
MONGO_URI = "mongodb+srv://maheshkumar17032k3:s4Lk5ve251iW6lml@yielddashboard.9wsy0ot.mongodb.net/?retryWrites=true&w=majority&appName=yielddashboard"
MAX_RETRIES = 5
RETRY_DELAY = 5

def get_mongo_client():
    """Initialize MongoDB client with TLS configuration and retries."""
    for attempt in range(MAX_RETRIES):
        try:
            client = MongoClient(
                MONGO_URI,
                tls=True,
                tlsAllowInvalidCertificates=True,
                connectTimeoutMS=30000,
                socketTimeoutMS=30000
            )
            client.admin.command('ping')
            db = client["yield_dashboard"]
            collection = db["yield_data"]
            logging.info("Successfully connected to MongoDB.")
            return client, db, collection
        except (ServerSelectionTimeoutError, AutoReconnect) as e:
            logging.error(f"Connection attempt {attempt + 1} failed: {e}")
            if attempt < MAX_RETRIES - 1:
                time.sleep(RETRY_DELAY * (2 ** attempt))
            else:
                logging.error(f"Max retries reached. Could not connect to MongoDB.")
                raise
        except Exception as e:
            logging.error(f"Unexpected error during connection attempt {attempt + 1}: {e}\n{traceback.format_exc()}")
            if attempt < MAX_RETRIES - 1:
                time.sleep(RETRY_DELAY * (2 ** attempt))
            else:
                raise
    return None, None, None

client, db, collection = get_mongo_client()

# Selenium setup for Microsoft Edge
edge_driver_path = r"C:\Users\ASUS\Downloads\edgedriver64\msedgedriver.exe"
service = Service(executable_path=edge_driver_path)
edge_options = Options()
# edge_options.add_argument("--headless")
driver = None

def close_excel_files():
    """Close all open Excel files to prevent access conflicts."""
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        for wb in excel.Workbooks:
            wb.Close(SaveChanges=False)
        excel.Quit()
        logging.info("Closed all open Excel files.")
    except Exception as e:
        logging.warning(f"Error closing Excel files: {e}\n{traceback.format_exc()}")

    for proc in psutil.process_iter(['name']):
        if proc.info['name'].lower() == 'excel.exe':
            proc.terminate()
            logging.info("Terminated Excel process.")

def find_header_row(sheet, header_text, column):
    """Find the row number where header_text appears in the specified column."""
    for row in range(1, sheet.max_row + 1):
        cell_value = sheet[f"{column}{row}"].value
        if cell_value and header_text.lower() in str(cell_value).lower():
            return row
    return None

def get_values_below_header(sheet, header_row, column):
    """Collect all non-empty values below the header row in the specified column."""
    values = []
    for row in range(header_row + 1, sheet.max_row + 1):
        cell_value = sheet[f"{column}{row}"].value
        if cell_value:
            values.append(str(cell_value))
    return values

def load_reference_mapping(ref_file, sheet_name):
    """Load error codes and root causes from the reference file using the specified sheet."""
    logging.info(f"Loading reference file: {ref_file}, sheet: {sheet_name}")
    if not os.path.exists(ref_file):
        logging.error(f"Reference file {ref_file} does not exist.")
        return {}, []
    try:
        ref_wb = openpyxl.load_workbook(ref_file)
        if sheet_name not in ref_wb.sheetnames:
            logging.error(f"Sheet {sheet_name} not found in {ref_file}. Available sheets: {ref_wb.sheetnames}")
            return {}, []
        ref_sheet = ref_wb[sheet_name]
        mapping = {}
        error_codes = []
        for row in ref_sheet.iter_rows(min_row=2, max_col=2, values_only=True):
            if row[0] and row[1]:
                mapping[str(row[0])] = str(row[1])
                error_codes.append(str(row[0]))
        logging.info(f"Loaded {len(mapping)} mappings from {ref_file}")
        return mapping, error_codes
    except Exception as e:
        logging.error(f"Error loading reference file {ref_file}: {e}\n{traceback.format_exc()}")
        return {}, []

def find_root_cause(value, mapping, error_codes):
    """Map error code to root cause with exact, partial, or minimum matching."""
    value_str = str(value).lower()
    for error_code in mapping:
        if value_str == error_code.lower():
            return mapping[error_code]
    for error_code in mapping:
        if re.search(re.escape(error_code), value_str, re.IGNORECASE):
            return mapping[error_code]
    value_words = set(re.split(r'_|\s+', value_str)) - {''}
    best_match = None
    best_score = 0
    for error_code in error_codes:
        error_words = set(re.split(r'_|\s+', error_code.lower())) - {''}
        common_words = value_words & error_words
        score = len(common_words)
        if score >= 1:
            if score > best_score:
                best_score = score
                best_match = error_code
            elif score == best_score and (best_match is None or len(error_code) < len(best_match)):
                best_match = error_code
    return mapping[best_match] if best_match else None

def process_line(input_file, ref_file, line_name, date, start_time_str, end_time_str):
    """Process an input file and return data for MongoDB."""
    logging.info(f"Processing input file: {input_file} for line: {line_name}")
    if not os.path.exists(input_file):
        logging.error(f"Input file {input_file} does not exist.")
        return None
    try:
        input_wb = openpyxl.load_workbook(input_file, data_only=True)
        if "Total" not in input_wb.sheetnames:
            logging.error(f"Sheet 'Total' not found in {input_file}. Available sheets: {input_wb.sheetnames}")
            return None
        total_sheet = input_wb["Total"]
    except Exception as e:
        logging.error(f"Error loading {input_file}: {e}\n{traceback.format_exc()}")
        return None

    input_value = total_sheet["V8"].value or 0
    logging.info(f"Input value from V8: {input_value}")
    header_row = find_header_row(total_sheet, "2nd Testcode", "U")
    if header_row is None:
        logging.error(f"'2nd Testcode' not found in {input_file}")
        return None

    testcodes = get_values_below_header(total_sheet, header_row, "U")
    if not testcodes:
        logging.warning(f"No test codes found in {input_file}")
        return None
    logging.info(f"Found {len(testcodes)} test codes in {input_file}")

    total_count = len(testcodes)
    unique_counts = Counter(testcodes)
    logging.info(f"Unique test codes: {len(unique_counts)}")

    # Use "Line 10" sheet for Line 01, Line 09, Line 10; otherwise, use line-specific sheet
    sheet_name = "Line 10" if line_name in ["Line 01", "Line 09", "Line 10"] else line_name
    mapping, error_codes = load_reference_mapping(ref_file, sheet_name)
    if not mapping:
        logging.error(f"No mappings loaded for {line_name}")
        return None

    root_cause_sums = Counter()
    unmatched_count = 0
    for testcode, count in unique_counts.items():
        root_cause = find_root_cause(testcode, mapping, error_codes)
        if root_cause:
            root_cause_sums[root_cause] += count
        else:
            unmatched_count += count

    total_failures = sum(root_cause_sums.values()) + unmatched_count
    yield_value = 1 - (total_failures / input_value) if input_value > 0 else 1
    logging.info(f"Calculated yield for {line_name}: {yield_value}, total failures: {total_failures}")

    tz = pytz.timezone('Asia/Kolkata')
    today = datetime.now(tz).date()
    start_dt = datetime.combine(today, datetime.strptime(start_time_str, "%H:%M:%S").time()).replace(tzinfo=tz)
    end_dt = datetime.combine(today, datetime.strptime(end_time_str, "%H:%M:%S").time()).replace(tzinfo=tz)
    if end_dt < start_dt:
        end_dt += timedelta(days=1)

    document = {
        "line": line_name,
        "date": start_dt.strftime("%Y-%m-%d"),
        "time_slot": f"{start_time_str[:5]}-{end_time_str[:5]}",
        "start_time": start_dt,
        "end_time": end_dt,
        "input": input_value,
        "yield": yield_value,
        "root_causes": [
            {"root_cause": rc, "count": count, "rate": count / input_value if input_value > 0 else 0}
            for rc, count in root_cause_sums.items()
        ],
        "other_failures": {"count": unmatched_count, "rate": unmatched_count / input_value if input_value > 0 else 0}
    }
    return document

def update_yield_report(input_dir, ref_dir, start_time, end_time):
    """Process all lines and insert data into MongoDB."""
    lines = [
        {"input_file": "Line 01.xlsx", "line_name": "Line 01"},
        {"input_file": "Line 03.xlsx", "line_name": "Line 03"},
        {"input_file": "Line 08.xlsx", "line_name": "Line 08"},
        {"input_file": "Line 09.xlsx", "line_name": "Line 09"},
        {"input_file": "Line 10.xlsx", "line_name": "Line 10"},
        {"input_file": "Line 12.xlsx", "line_name": "Line 12"},
        {"input_file": "Line 15.xlsx", "line_name": "Line 15"},
    ]
    ref_files = {
        "Line 01": "Reference.xlsx",
        "Line 09": "Reference.xlsx",
        "Line 10": "Reference.xlsx",
        "Line 03": "Kansas.xlsx",
        "Line 08": "Manila.xlsx",
        "Line 12": "Lamulite.xlsx",
        "Line 15": "Vegas.xlsx",
    }
    date = datetime.now(pytz.timezone('Asia/Kolkata')).strftime("%Y-%m-%d")
    
    close_excel_files()
    
    for line in lines:
        logging.info(f"Starting processing for {line['line_name']}")
        input_file = os.path.join(input_dir, line["input_file"])
        ref_file = os.path.join(ref_dir, ref_files[line["line_name"]])
        document = process_line(input_file, ref_file, line["line_name"], date, start_time, end_time)
        if document:
            try:
                collection.insert_one(document)
                logging.info(f"Inserted data for {line['line_name']} into MongoDB")
            except Exception as e:
                logging.error(f"Failed to insert data for {line['line_name']}: {e}\n{traceback.format_exc()}")
        else:
            logging.warning(f"No document generated for {line['line_name']}")

def initialize_driver():
    """Initialize the WebDriver."""
    global driver
    if driver is not None:
        try:
            driver.quit()
        except:
            pass
    driver = webdriver.Edge(service=service, options=edge_options)
    driver.maximize_window()
    logging.info("WebDriver initialized")
    return driver

def set_dates_and_times(start_time, end_time):
    """Set date and time fields on the website."""
    logging.info(f"Setting dates and times: start={start_time}, end={end_time}")
    url = "https://mqs.motorola.com/NPI/NTF_Pareto_Split.aspx?enc=KAvT2iht37lGXVuV1DDr6m2J4xZhw9rxhhrM63UWwSHkmWErJ/8zW8C9qT/DS5xhjnRJXpPUEbr3iEwvrSqeKgrpCO8KeuYi7vM3CtESTrEMV6B0bA4409qm96+Ft/cI"
    driver.get(url)
    wait = WebDriverWait(driver, 15)
    
    try:
        start_date_calendar = wait.until(EC.element_to_be_clickable((By.ID, "Accordion_Normal_content_Image1")))
        start_date_calendar.click()
        today_button = wait.until(EC.element_to_be_clickable((By.ID, "Accordion_Normal_content_calendarButtonExtender_today")))
        today_button.click()
        
        end_date_calendar = wait.until(EC.element_to_be_clickable((By.ID, "Accordion_Normal_content_Image2")))
        end_date_calendar.click()
        today_button = wait.until(EC.element_to_be_clickable((By.ID, "Accordion_Normal_content_CalendarExtender1_today")))
        today_button.click()
        
        start_time_input = wait.until(EC.presence_of_element_located((By.ID, "Accordion_Normal_content_NewTxtStartTime")))
        start_time_input.clear()
        start_time_input.send_keys(start_time)
        
        end_time_input = wait.until(EC.presence_of_element_located((By.ID, "Accordion_Normal_content_NewTxtEndTime")))
        end_time_input.clear()
        end_time_input.send_keys(end_time)
        logging.info("Successfully set dates and times")
    except Exception as e:
        logging.error(f"Error setting dates and times: {e}\n{traceback.format_exc()}")
        raise

def select_line_and_generate_report(line):
    """Select a line and generate its report."""
    logging.info(f"Selecting line: {line}")
    wait = WebDriverWait(driver, 15)
    try:
        specific_radio = wait.until(EC.element_to_be_clickable((By.ID, "TabContainer2_TabPanel3_LineSpec")))
        if not specific_radio.is_selected():
            specific_radio.click()
        
        checkboxes = driver.find_elements(By.XPATH, "//input[@type='checkbox' and contains(@id, 'LineList')]")
        for checkbox in checkboxes:
            if checkbox.is_selected():
                checkbox.click()
        
        if line == "BE01":
            checkbox_id, file_name = "TabContainer2_TabPanel3_LineList_14", "Line 01.xlsx"
        elif line == "BE03":
            checkbox_id, file_name = "TabContainer2_TabPanel3_LineList_18", "Line 03.xlsx"
        elif line == "BE08":
            checkbox_id, file_name = "TabContainer2_TabPanel3_LineList_23", "Line 08.xlsx"
        elif line == "BE09":
            checkbox_id, file_name = "TabContainer2_TabPanel3_LineList_24", "Line 09.xlsx"
        elif line == "BE10":
            checkbox_id, file_name = "TabContainer2_TabPanel3_LineList_25", "Line 10.xlsx"
        elif line == "BE12":
            checkbox_id, file_name = "TabContainer2_TabPanel3_LineList_27", "Line 12.xlsx"
        elif line == "BE15":
            checkbox_id, file_name = "TabContainer2_TabPanel3_LineList_30", "Line 15.xlsx"
        else:
            raise ValueError(f"Unknown line: {line}")
        
        checkbox = wait.until(EC.element_to_be_clickable((By.ID, checkbox_id)))
        checkbox.click()
        generate_button = wait.until(EC.element_to_be_clickable((By.ID, "ButGenerateReport")))
        generate_button.click()
        time.sleep(3)
        logging.info(f"Report generated for {line}, file: {file_name}")
        return file_name
    except Exception as e:
        logging.error(f"Error selecting line {line}: {e}\n{traceback.format_exc()}")
        raise

def handle_downloaded_file(file_name):
    """Move downloaded file to REPORT_DIR."""
    logging.info(f"Handling downloaded file: {file_name}")
    downloads_dir = os.path.join(os.path.expanduser("~"), "Downloads")
    try:
        latest_file = max([os.path.join(downloads_dir, f) for f in os.listdir(downloads_dir)], key=os.path.getctime)
        destination_path = os.path.join(REPORT_DIR, file_name)
        if os.path.exists(destination_path):
            os.remove(destination_path)
        shutil.move(latest_file, destination_path)
        logging.info(f"Moved {latest_file} to {destination_path}")
    except Exception as e:
        logging.error(f"Error handling downloaded file {file_name}: {e}\n{traceback.format_exc()}")
        raise

def get_time_range():
    """Calculate start, end, and execution times."""
    now = datetime.now()
    if now.minute > 30 or (now.minute == 30 and now.second > 0):
        end_time_dt = now.replace(minute=30, second=0, microsecond=0)
    else:
        end_time_dt = now.replace(minute=30, second=0, microsecond=0) - timedelta(hours=1)
    start_time_dt = end_time_dt - timedelta(hours=1)
    execution_time_dt = end_time_dt + timedelta(minutes=5)
    return (start_time_dt.strftime("%H:%M:%S"), end_time_dt.strftime("%H:%M:%S"), execution_time_dt)

def main():
    """Run the scraper and update MongoDB hourly."""
    global driver, client, db, collection
    while True:
        try:
            if client is None or not client.server_info:
                client, db, collection = get_mongo_client()
                if client is None:
                    raise ServerSelectionTimeoutError("Failed to establish MongoDB connection after retries.")

            driver = initialize_driver()
            start_time, end_time, execution_time_dt = get_time_range()
            sleep_time = (execution_time_dt - datetime.now()).total_seconds()
            if sleep_time > 0:
                logging.info(f"Waiting {sleep_time:.0f}s until {execution_time_dt.strftime('%H:%M:%S')}")
                time.sleep(sleep_time)
            
            logging.info(f"Processing data from {start_time} to {end_time}")
            set_dates_and_times(start_time, end_time)
            
            for line in ["BE01", "BE03", "BE08", "BE09", "BE10", "BE12", "BE15"]:
                logging.info(f"Processing line: {line}")
                file_name = select_line_and_generate_report(line)
                handle_downloaded_file(file_name)
            
            update_yield_report(REPORT_DIR, REF_DIR, start_time, end_time)
            next_execution = execution_time_dt + timedelta(hours=1)
            sleep_time = (next_execution - datetime.now()).total_seconds()
            if sleep_time > 0:
                time.sleep(sleep_time)
        except WebDriverException as e:
            logging.error(f"WebDriver error (possibly browser closed): {e}\n{traceback.format_exc()}")
            if driver:
                try:
                    driver.quit()
                except:
                    pass
            driver = None
            time.sleep(5)
            continue
        except (ServerSelectionTimeoutError, AutoReconnect) as e:
            logging.error(f"MongoDB connection error: {e}\n{traceback.format_exc()}")
            client, db, collection = None, None, None
            time.sleep(5)
            continue
        except Exception as e:
            logging.error(f"Unexpected error: {e}\n{traceback.format_exc()}")
            if driver:
                try:
                    driver.quit()
                except:
                    pass
            driver = None
            time.sleep(5)
            continue
        finally:
            pass

if __name__ == "__main__":
    main()